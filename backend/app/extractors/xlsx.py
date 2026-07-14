import io

from openpyxl import load_workbook

MAX_ROWS_PER_SHEET = 500


def extract(content: bytes) -> tuple[str, list[str]]:
    warnings: list[str] = []
    # data_only: se leen los valores calculados, nunca se evalúan fórmulas.
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    parts: list[str] = []

    for name in wb.sheetnames:
        ws = wb[name]
        rows: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= MAX_ROWS_PER_SHEET:
                warnings.append(f"Hoja '{name}' truncada en {MAX_ROWS_PER_SHEET} filas")
                break
            cells = ["" if c is None else str(c) for c in row]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            parts.append(f"## Hoja: {name}\n" + "\n".join(rows))

    wb.close()
    return "\n\n".join(parts), warnings


def metadata(content: bytes) -> dict:
    wb = load_workbook(io.BytesIO(content), read_only=True)
    props = wb.properties
    meta = {
        "author": props.creator,
        "title": props.title,
        "created": props.created.isoformat() if props.created else None,
        "modified": props.modified.isoformat() if props.modified else None,
        "last_modified_by": props.lastModifiedBy,
        "sheets": list(wb.sheetnames),
    }
    wb.close()
    return meta
